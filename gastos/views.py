import openpyxl
import os
import io
from PIL import Image
from .sharepoint import respaldar_en_sharepoint
from openpyxl.styles import Font, PatternFill, Border, Side
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.contrib import messages
from datetime import date, datetime 
from django.urls import reverse
from django.utils.http import urlencode
from django.http import HttpResponse, Http404
from django.db.models.functions import TruncMonth
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from .models import Gasto, Empresa, Obra, Presupuesto
from .forms import GastoForm, CargaMasivaForm,ObraForm
from .utils import procesar_boleta_chilena

def get_gastos_empresa(user):
    # Si es superusuario, ve todo (opcional)
    if user.is_superuser:
        return Gasto.objects.all().order_by('-fecha_emision')
    
    # Si es mortal, verificamos si tiene empresa asignada
    if hasattr(user, 'perfil') and user.perfil.empresa:
        return Gasto.objects.filter(empresa=user.perfil.empresa).order_by('-fecha_emision')
    
    # Si no tiene empresa, no ve nada
    return Gasto.objects.none()

@login_required
def dashboard(request):
    gastos = get_gastos_empresa(request.user)

    obras = []
    obra_seleccionada = request.GET.get('obra') 
    
    if hasattr(request.user, 'perfil') and request.user.perfil.empresa:
        obras = Obra.objects.filter(empresa=request.user.perfil.empresa, activo=True)
        if obra_seleccionada:
            gastos = gastos.filter(obra_id=obra_seleccionada)

    total_mes = 0
    hoy = date.today()
    total_validado = gastos.aggregate(Sum('monto_total'))['monto_total__sum'] or 0
    cantidad_gastos = gastos.count()

    # --- 1. NUEVOS CÁLCULOS: PRESUPUESTO Y DISPONIBLE ---
    presupuesto_total = 0
    
    # Si hay una obra seleccionada, sumamos los presupuestos de esa obra
    if obra_seleccionada:
        presupuesto_total = Presupuesto.objects.filter(obra_id=obra_seleccionada).aggregate(Sum('monto'))['monto__sum'] or 0
    # Si estamos viendo "Todas las obras", sumamos todos los presupuestos de la empresa
    elif obras:
        presupuesto_total = Presupuesto.objects.filter(obra__in=obras).aggregate(Sum('monto'))['monto__sum'] or 0

    presupuesto_disponible = presupuesto_total - total_validado
    
    porcentaje_gastado = 0
    if presupuesto_total > 0:
        porcentaje_gastado = (total_validado / presupuesto_total) * 100
        porcentaje_gastado = round(porcentaje_gastado, 1) # Redondeamos a 1 decimal

    # --- 2. DATOS PARA GRÁFICO DE LÍNEA (Evolución) ---
    gastos_por_mes = (
        gastos
        .annotate(mes=TruncMonth('fecha_emision'))
        .values('mes')
        .annotate(total=Sum('monto_total'))
        .order_by('mes')
    )
    
    labels_grafico = []
    data_grafico = []
    for g in gastos_por_mes:
        if g['mes']:
            fecha_fmt = g['mes'].strftime("%Y-%m")
            labels_grafico.append(fecha_fmt)
            data_grafico.append(g['total'])

    # --- 3. DATOS PARA GRÁFICO DE BARRAS Y DONA (Categorías) ---
    resumen_categorias = (
        gastos
        .values('categoria')
        .annotate(total=Sum('monto_total'))
        .order_by('categoria')
    )
    gastos_dict = {item['categoria']: item['total'] for item in resumen_categorias}
    
    presupuesto_dict = {}
    if obra_seleccionada:
        presupuestos = Presupuesto.objects.filter(obra_id=obra_seleccionada)
        presupuesto_dict = {p.categoria: p.monto for p in presupuestos}
    
    labels_cat = []
    data_cat = []
    data_presup = []
    #nombres_legibles = dict(Gasto.CATEGORIAS)

    for codigo, nombre in Gasto.CATEGORIAS:
        real = gastos_dict.get(codigo, 0)
        meta = presupuesto_dict.get(codigo, 0)

        if real > 0 or meta > 0:
            labels_cat.append(nombre)
            data_cat.append(real)
            data_presup.append(meta)

    # --- 4. ÚLTIMOS MOVIMIENTOS (Tabla) ---
    ultimos_gastos = gastos.order_by('-fecha_emision', '-id')[:5]

    # --- 5. ENVIAR TODO AL HTML ---
    context = {
        'total_mes': total_mes,
        'cantidad_gastos': cantidad_gastos,
        'total_validado': total_validado,
        
        # Nuevas variables enviadas al HTML:
        'presupuesto_total': presupuesto_total,
        'presupuesto_disponible': presupuesto_disponible,
        'porcentaje_gastado': porcentaje_gastado,

        'labels_grafico': labels_grafico,
        'data_grafico': data_grafico,
        'ultimos_gastos': ultimos_gastos,
        'labels_cat': labels_cat,
        'data_cat': data_cat,
        'data_presup': data_presup,
        'fecha_actual': hoy,
        'obras': obras,                        
        'obra_seleccionada': obra_seleccionada 
    }

    return render(request, 'gastos/dashboard.html', context)


@login_required
def crear_gasto(request):
    if request.method == 'POST':
        # Pasamos 'request.user' al formulario para filtrar el dropdown de Obras
        form = GastoForm(request.user, request.POST, request.FILES)
        
        if form.is_valid():
            gasto = form.save(commit=False)
            gasto.usuario = request.user 

            # Asignamos la Empresa automáticamente
            if hasattr(request.user, 'perfil') and request.user.perfil.empresa:
                gasto.empresa = request.user.perfil.empresa

            gasto.save() # Se guarda la imagen físicamente en el servidor

            # --- 1. LÓGICA DE ESCÁNER (OCR) ---
            if gasto.imagen: 
                try:
                    resultado = procesar_boleta_chilena(gasto.imagen.path)
                    
                    if 'error' not in resultado:
                        if resultado.get('monto_total'):
                            gasto.monto_total = resultado['monto_total']
                        
                        if resultado.get('rut_emisor'):
                            gasto.rut_emisor = resultado['rut_emisor']

                        if resultado.get('folio'):
                            gasto.folio = resultado['folio']

                        fecha_str = resultado.get('fecha_emision')
                        if fecha_str:
                            try:
                                gasto.fecha_emision = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                            except ValueError:
                                pass

                        gasto.procesado_exitosamente = True
                        gasto.save() # Se guardan los datos extraídos
                        messages.success(request, "¡Boleta leída! Por favor confirma los datos.")
                    else:
                        messages.warning(request, f"Escáner: {resultado['error']}")
                except Exception as e:
                    messages.warning(request, f"Error procesando imagen: {e}")
            
            # --- 2. RESPALDO EMPRESARIAL (SHAREPOINT) ---
            # Mandamos una copia a SharePoint en segundo plano
            respaldar_en_sharepoint(gasto)
            # ---------------------------------------------

            return redirect('editar_gasto', pk=gasto.pk)
    else:
        # Pasamos 'request.user' en el GET para que el formulario se inicie correctamente
        form = GastoForm(request.user)

    return render(request, 'gastos/crear_gasto.html', {'form': form, 'titulo': 'Nuevo Gasto'})

@login_required
def editar_gasto(request, pk):
    # Buscamos el gasto que el usuario quiere editar
    gasto = get_object_or_404(Gasto, pk=pk)
    
    if request.method == 'POST':
        # Pasamos instance=gasto para decirle a Django que estamos editando, no creando
        form = GastoForm(request.user, request.POST, request.FILES, instance=gasto)
        
        if form.is_valid():
            gasto = form.save()

            # --- MAGIA DE SINCRONIZACIÓN ---
            # form.changed_data es una lista que contiene los nombres de los campos que el usuario modificó
            if 'imagen' in form.changed_data or 'obra' in form.changed_data:
                
                print(f"🔄 [EDICIÓN] Detectado cambio en imagen u obra. Resincronizando Gasto #{gasto.id}...")
                
                # Volvemos a llamar a la función.
                # - Si cambió la foto pero no el proyecto: Microsoft sobrescribe la foto vieja.
                # - Si cambió el proyecto: Microsoft guarda la foto en la nueva carpeta.
                respaldar_en_sharepoint(gasto)
                
            # -------------------------------

            messages.success(request, "¡Gasto actualizado y sincronizado correctamente!")
            return redirect('historial_gastos') # O la ruta a la que redirijas normalmente
    else:
        form = GastoForm(request.user, instance=gasto)

    return render(request, 'gastos/editar_gasto.html', {
        'form': form, 
        'titulo': f'Editar Gasto #{gasto.folio or gasto.id}',
        'gasto': gasto
    })

@login_required
def eliminar_gasto(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk)
    if request.method == 'POST':
        gasto.delete()
        messages.success(request, "Gasto eliminado.")
        return redirect('dashboard')
    return render(request, 'gastos/eliminar_gasto.html', {'object': gasto})

@login_required
def historial_gastos(request):
    # 1. Obtener la empresa del usuario
    empresa_usuario = None
    if hasattr(request.user, 'perfil') and request.user.perfil.empresa:
        empresa_usuario = request.user.perfil.empresa

    # 2. Base de datos: Traer solo gastos de esa empresa
    if empresa_usuario:
        gastos = Gasto.objects.filter(empresa=empresa_usuario)
        obras = Obra.objects.filter(empresa=empresa_usuario, activo=True)
        categorias_brutas = Gasto.objects.filter(empresa=empresa_usuario).exclude(categoria__isnull=True).exclude(categoria__exact='').values_list('categoria', flat=True).distinct()
        categorias_disponibles = []
        for cat in categorias_brutas:
            nombre_bonito = cat.replace('_', ' ').title().replace(' De ', ' de ')
            categorias_disponibles.append({
                'valor_db': cat,
                'nombre_visible': nombre_bonito
            })
    else:
        gastos = Gasto.objects.none()
        obras = Obra.objects.none()
        categorias_disponibles = []
    
    

    # 3. Capturar lo que el usuario escribió en los filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    obra_id = request.GET.get('obra_id') # Capturamos el nuevo filtro
    categoria_seleccionada = request.GET.get('categoria') # NUEVO

    # 4. Aplicar los filtros si existen
    if fecha_inicio:
        gastos = gastos.filter(fecha_emision__gte=fecha_inicio)
    
    if fecha_fin:
        gastos = gastos.filter(fecha_emision__lte=fecha_fin)
        
    if obra_id:
        gastos = gastos.filter(obra_id=obra_id) # Filtramos por el proyecto seleccionado
    
    if categoria_seleccionada:
        gastos = gastos.filter(categoria=categoria_seleccionada)

    # 5. Ordenar (los más recientes primero) y calcular el total filtrado
    gastos = gastos.order_by('-fecha_emision', '-id')
    total_filtrado = gastos.aggregate(total=Sum('monto_total'))['total'] or 0

    # 6. Enviar todo al HTML
    context = {
        'gastos': gastos,
        'obras': obras, # Mandamos la lista de proyectos al select
        'categorias_disponibles': categorias_disponibles,
        'total_filtrado': total_filtrado,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'obra_seleccionada': obra_id, # Mandamos el ID seleccionado para que no se borre al recargar
        'categoria_seleccionada': categoria_seleccionada,
    }
    
    return render(request, 'gastos/historial.html', context)

@login_required
def carga_masiva(request):
    if request.method == 'POST':
        print("DEBUG: --- Iniciando Carga Masiva ---")
        
        form = CargaMasivaForm(request.POST, request.FILES)
        
        if form.is_valid():
            # USAMOS getlist PARA OBTENER TODOS LOS ARCHIVOS
            archivos = request.FILES.getlist('imagenes')
            print(f"DEBUG: Se recibieron {len(archivos)} archivos.")
            
            procesados = 0
            fallidos = 0
            
            for f in archivos:
                try:
                    print(f"DEBUG: Procesando archivo: {f.name}")
                    
                    # 1. Crear Gasto Base
                    nuevo_gasto = Gasto(
                        usuario=request.user,
                        imagen=f,
                        procesado_exitosamente=False
                    )
                    nuevo_gasto.save()
                    
                    # 2. Escanear
                    resultado = procesar_boleta_chilena(nuevo_gasto.imagen.path)
                    
                    if 'error' not in resultado:
                        if resultado.get('monto_total'):
                            nuevo_gasto.monto_total = resultado['monto_total']
                        if resultado.get('rut_emisor'):
                            nuevo_gasto.rut_emisor = resultado['rut_emisor']
                        if resultado.get('folio'):
                            nuevo_gasto.folio = resultado['folio']
                        
                        # Fecha
                        fecha_str = resultado.get('fecha_emision')
                        if fecha_str:
                            try:
                                nuevo_gasto.fecha_emision = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                            except ValueError: pass
                        
                        nuevo_gasto.procesado_exitosamente = True
                        nuevo_gasto.save()
                        procesados += 1
                    else:
                        print(f"DEBUG: Falló lectura de {f.name}: {resultado['error']}")
                        fallidos += 1
                        
                except Exception as e:
                    print(f"DEBUG: Error crítico en {f.name}: {e}")
                    fallidos += 1

            # Mensajes al usuario
            if procesados > 0:
                messages.success(request, f"¡Éxito! {procesados} boletas procesadas correctamente.")
            if fallidos > 0:
                messages.warning(request, f"{fallidos} boletas no se pudieron leer. Revísalas en el historial.")
            
            return redirect('historial_gastos')
        
        else:
            print("DEBUG: Formulario inválido. Errores:")
            print(form.errors)
            messages.error(request, "Error en el formulario. Revisa la consola.")
    
    else:
        form = CargaMasivaForm()

    return render(request, 'gastos/carga_masiva.html', {'form': form})

@login_required
def eliminar_masivo(request):
    if request.method == 'POST':
        ids_a_borrar = request.POST.getlist('gastos_ids')
        
        # --- Lógica de Borrado (Igual que antes) ---
        if ids_a_borrar:
            cantidad, _ = Gasto.objects.filter(
                id__in=ids_a_borrar, 
                usuario=request.user
            ).delete()
            
            if cantidad > 0:
                messages.success(request, f"🗑️ Se eliminaron {cantidad} boletas.")
            else:
                messages.warning(request, "No se pudo eliminar.")
        
        # --- NUEVA LÓGICA DE REDIRECCIÓN INTELIGENTE ---
        
        # 1. Recuperamos los filtros que venían ocultos en el formulario
        f_inicio = request.POST.get('fecha_inicio_filtro')
        f_fin = request.POST.get('fecha_fin_filtro')
        
        # 2. Obtenemos la URL base del historial (ej: '/gastos/historial/')
        base_url = reverse('historial_gastos')
        
        # 3. Preparamos los parámetros (query string)
        parametros = {}
        if f_inicio: parametros['fecha_inicio'] = f_inicio
        if f_fin:    parametros['fecha_fin'] = f_fin
        
        # 4. Si hay parámetros, los pegamos a la URL
        if parametros:
            # Esto crea algo como: /gastos/historial/?fecha_inicio=2024-01-01&fecha_fin=...
            url_final = f"{base_url}?{urlencode(parametros)}"
            return redirect(url_final)
            
    # Si no hay filtros, volvemos al historial normal
    return redirect('historial_gastos')

@login_required
def exportar_excel(request):
    # 1. Recuperar Filtros (Igual que en el historial)
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    gastos = Gasto.objects.all().order_by('-fecha_emision')
    if fecha_inicio: gastos = gastos.filter(fecha_emision__gte=fecha_inicio)
    if fecha_fin:    gastos = gastos.filter(fecha_emision__lte=fecha_fin)

    # 2. Configurar el Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Reporte_Gastos_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"

    # 3. Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # 4. Cabecera
    headers = ["Fecha", "RUT Emisor", "Folio", "Monto Total", "Usuario"]
    ws.append(headers)
    
    for cell in ws[1]: # Aplicar estilo a la primera fila
        cell.font = header_font
        cell.fill = header_fill

    # 5. Datos
    total = 0
    for gasto in gastos:
        ws.append([
            gasto.fecha_emision,
            gasto.rut_emisor,
            gasto.folio,
            gasto.monto_total,
            gasto.usuario.username
        ])
        total += gasto.monto_total

    # 6. Fila de Total
    ws.append(["", "", "TOTAL ACUMULADO:", total, ""])
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)

    wb.save(response)
    return response

@login_required
def exportar_pdf(request):
    # 1. Recuperar Filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    gastos = Gasto.objects.all().order_by('-fecha_emision')
    if fecha_inicio: gastos = gastos.filter(fecha_emision__gte=fecha_inicio)
    if fecha_fin:    gastos = gastos.filter(fecha_emision__lte=fecha_fin)
    
    total = gastos.aggregate(Sum('monto_total'))['monto_total__sum'] or 0

    # 2. Renderizar HTML para el PDF
    template_path = 'gastos/reporte_pdf.html'
    context = {
        'gastos': gastos,
        'total': total,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'empresa': 'Mi Constructora S.A.' # Puedes personalizar esto
    }
    
    # 3. Crear PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_gastos.pdf"'
    
    html = render_to_string(template_path, context)
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Tuvimos errores generando el PDF <pre>' + html + '</pre>')
    return response

@login_required
def lista_obras(request):
    # Solo traemos las obras de MI empresa
    if hasattr(request.user, 'perfil') and request.user.perfil.empresa:
        obras = Obra.objects.filter(empresa=request.user.perfil.empresa)
    else:
        obras = []
    return render(request, 'gastos/lista_obras.html', {'obras': obras})

# --- CREAR OBRA NUEVA ---
@login_required
def crear_obra(request):
    if request.method == 'POST':
        form = ObraForm(request.POST)
        if form.is_valid():
            obra = form.save(commit=False)
            # Asignamos automáticamente la empresa del usuario
            if hasattr(request.user, 'perfil') and request.user.perfil.empresa:
                obra.empresa = request.user.perfil.empresa
                obra.save()
                return redirect('lista_obras') # Volver a la lista
    else:
        form = ObraForm()
    
    return render(request, 'gastos/crear_obra.html', {'form': form})

def nuevo_gasto(request):
    if request.method == 'POST':
        form = GastoForm(request.POST, request.FILES)
        if form.is_valid():
            gasto = form.save(commit=False)
            
            # --- ASIGNACIÓN AUTOMÁTICA ---
            gasto.usuario = request.user
            if hasattr(request.user, 'perfil'):
                gasto.empresa = request.user.perfil.empresa
            # -----------------------------
            
            gasto.save()
            return redirect('dashboard')
    else:
        form = GastoForm()
    return render(request, 'gastos/nuevo_gasto.html', {'form': form})

@login_required
def definir_presupuesto(request, obra_id):
    # 1. Buscamos la obra y verificamos seguridad (que sea de mi empresa)
    obra = get_object_or_404(Obra, pk=obra_id)
    if not (hasattr(request.user, 'perfil') and request.user.perfil.empresa == obra.empresa):
        messages.error(request, "No tienes permiso para ver esta obra.")
        return redirect('lista_obras')

    # 2. Obtenemos las categorías disponibles en el sistema
    categorias = Gasto.CATEGORIAS

    if request.method == 'POST':
        # 3. Guardamos los datos del formulario manual
        for codigo, nombre in categorias:
            # Buscamos el input con el nombre 'presupuesto_MATERIALES', etc.
            monto_str = request.POST.get(f'presupuesto_{codigo}')
            
            if monto_str:
                try:
                    monto = int(monto_str)
                    # update_or_create: Si existe lo actualiza, si no, lo crea
                    Presupuesto.objects.update_or_create(
                        obra=obra,
                        categoria=codigo,
                        defaults={'monto': monto}
                    )
                except ValueError:
                    pass # Si escribieron texto en vez de número, lo ignoramos
        
        messages.success(request, "¡Presupuesto actualizado correctamente!")
        return redirect('lista_obras')

    # 4. Preparamos los datos para mostrar en el HTML (Montos actuales)
    # Creamos un diccionario para acceso rápido: {'MATERIALES': 500000, ...}
    presupuestos_actuales = {
        p.categoria: p.monto 
        for p in Presupuesto.objects.filter(obra=obra)
    }

    datos_tabla = []
    for codigo, nombre in categorias:
        datos_tabla.append({
            'codigo': codigo,
            'nombre': nombre,
            'monto_actual': presupuestos_actuales.get(codigo, 0) # Si no existe, es 0
        })

    return render(request, 'gastos/definir_presupuesto.html', {
        'obra': obra,
        'datos_tabla': datos_tabla
    })

@login_required
def descargar_boleta_pdf(request, pk):
    """
    Descarga el archivo adjunto. Si es una imagen, la convierte a PDF al vuelo.
    """
    gasto = get_object_or_404(Gasto, pk=pk)

    if not gasto.imagen or not hasattr(gasto.imagen, 'path'):
        raise Http404("Este gasto no tiene ningún archivo adjunto.")

    ruta_archivo = gasto.imagen.path
    extension = os.path.splitext(ruta_archivo)[1].lower()
    
    # Armamos un nombre profesional para el archivo descargado
    nombre_descarga = f"Boleta_{gasto.folio or 'SF'}_{gasto.rut_emisor or 'SinRut'}.pdf"

    # CASO A: Ya es un PDF. Lo devolvemos directamente.
    if extension == '.pdf':
        with open(ruta_archivo, 'rb') as pdf:
            response = HttpResponse(pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{nombre_descarga}"'
            return response

    # CASO B: Es una imagen. La convertimos a PDF.
    elif extension in ['.jpg', '.jpeg', '.png']:
        try:
            imagen = Image.open(ruta_archivo)
            
            # Si es un PNG con transparencia, lo pasamos a formato estándar (RGB) para que el PDF no falle
            if imagen.mode in ("RGBA", "P"):
                imagen = imagen.convert("RGB")

            # Creamos un "archivo falso" en la memoria de Python
            buffer = io.BytesIO()
            # Guardamos la imagen en ese archivo como PDF
            imagen.save(buffer, format='PDF', resolution=100.0)
            buffer.seek(0)

            # Se lo enviamos al usuario
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{nombre_descarga}"'
            return response
            
        except Exception as e:
            return HttpResponse(f"Error convirtiendo la imagen a PDF: {str(e)}", status=500)
            
    # CASO C: Formato raro
    else:
        return HttpResponse("El formato del archivo adjunto no está soportado.", status=400)

